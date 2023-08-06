import json, copy
from openpyxl import load_workbook, Workbook
from openpyxl.workbook.views import BookView
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from base.source.infosheet import SheetDict
from base.source.tablesheet import TableList
import warnings
from base.source.excelwriter import ExcelWritter
from base.utils.utils import DateEncoder

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class Excel(object):
    """Excel manupulate excel for getting source data, compare excels or sheets data, and write new excel for collecting data.

    Args:
        excel_name ([str]): [an excel file name]
        sheets ([list]): [default is None, meaning will get all sheets; If assing a list of sheet names, it will only access the specified sheets.
        sheet name in this Excel class will with its original 'info-' or 'table-' head, but the head will be removed when save to dict for exporting;
        When write to a new excel file, the sheet name of a object of SheetDict will be insert 'info-' in the name, while for TableList object, 'table-'
        will be inserted too]
    Returns: an Excel object
    """

    def __init__(self, excel_name, sheets=None, language="English"):
        self.language = language
        self.excel_name = excel_name
        self.wb = load_workbook(self.excel_name)
        self.sheetNames = (
            self.wb.sheetnames if sheets == None else sheets
        )  # if not specify sheets, it will handle all sheets
        self.sheet_names = [
            sheetName
            for sheetName in self.sheetNames
            if sheetName.lower().startswith("info-")
        ]  # save info sheet names
        self.table_names = [
            sheetName
            for sheetName in self.sheetNames
            if sheetName.lower().startswith("table-")
        ]  # save table sheet names
        self.sheets = OrderedDict()  # save info sheet objects
        self.tables = OrderedDict()  # save table sheet objects
        self._getSheetsTables()  # get all info and table sheets objects and save to sheets and tables

    def __str__(self):
        return (
            self.excel_name
            + " with info sheets: "
            + ", ".join(self.sheets)
            + " , and table sheets: "
            + " ,".join(self.tables)
        )

    # = operator. Return True if self and another excel have same sheets, and every sheet has same variables
    def __eq__(self, another: object):
        if self.sheetNames != another.sheetNames:
            return False  # if not all sheet names same, False
        # check if info sheets are same
        pairs = list(zip(self.sheets, another.sheets))
        for self_pair, another_pair in pairs:
            if self_pair != another.pair:
                return False
        # check if table sheets are same
        pairs = list(zip(self.tables, another.tables))
        for self_pair, another_pair in pairs:
            if self_pair != another.pair:
                return False
        return True

    # + union two excel objs (self | another), and return a new excel obj
    # in excel level, I consider to merge value together with variables. If duplicated, another value will prevail.
    def __add__(self, another: object):
        # iterate another and insert or replace self's
        # merge sheets (info and talbe)
        for sheet in another.sheets:
            if sheet not in self.sheets:  # sheet is a key of sheets OrderedDict
                self.sheets[sheet] = another.sheets[sheet]
            else:
                self.sheets[sheet] += another.sheets[sheet]
        for table in another.tables:
            if table not in self.tables:
                # 创建一个新的table
                self.tables[table] = another.tables[table]
            else:
                self.tables[table] += another.tables[
                    table
                ]  # 原来是每个子项都自己算，v1.0.3不做了。 从V1.0.4改回来。
                # append rows to old one; 直接把后面的excel table里面的rows添加到签前面的table里面。
                # self.tables[table].data.extend(another.tables[table].data) # 这个是添加到后面的版本。 只有在V1.0.3里面用。

        self.sheet_names += another.sheet_names
        self.table_names += another.table_names
        return self

    # - set difference, return a new obj with sheets in self but not in another, and with variables in self sheet but not in another if has same sheet
    def __sub__(self, another: object):
        # for sheet in obj.sheets:
        for sheet in self.sheets:
            if (
                sheet in another.sheets
            ):  # if sheet in obj and in another, self sheet minus another sheet. sheet is a key of sheets OrderedDict
                self.sheets[sheet] -= another.sheets[sheet]
            # if len(self.sheets[sheet])==0:   # if a sheet is empty, delete it
            #     self.sheets.pop(sheet)

        # table part: for columns sub, only care about variables; while for rows, care about if same row
        # for table in self.tables:
        for table in self.tables:
            if (
                table in another.tables
            ):  # if table in obj and in another, self sheet minus another sheet. sheet is a key of sheets OrderedDict
                self.tables[table] -= another.tables[table]
            # if len(self.tables[table])==0:
            #     self.tables.pop(table)
        self.sheet_names = list(set(self.sheet_names) - set(another.sheet_names))
        self.table_names = list(set(self.table_names) - set(another.table_names))

        return self

    # copy another's excel's common contents to self
    def copy(self, another: object):
        # iterate another's sheet, replace self's value if it is in self too
        for sheet in another.sheets:
            if sheet in self.sheets:  # sheet is a key of sheets OrderedDict
                self.sheets[sheet] = self.sheets[sheet].copy(another.sheets[sheet])
        for table in another.tables:
            if table in self.tables:
                self.tables[table] = self.tables[table].copy(another.tables[table])
        return self

    """
    Below are methods for getting data from excel files and export 
    """
    # get sheet variable name from sheet name
    def _getVariable(self, talbe_sheet_name):
        if talbe_sheet_name.startswith("info-"):
            return talbe_sheet_name[5:]
        elif talbe_sheet_name.startswith("table-"):
            return talbe_sheet_name[6:]
        else:
            return talbe_sheet_name

    # check if a row is None. Return True if there is at least one column not None
    def _notNone(self, row_data):
        return any([False if x == None else True for x in row_data])

    # get info sheet object
    def _getSheet(self, sheet):
        values = list(self.wb[sheet].values)
        values = [x for x in values if self._notNone(x)]  # clean data by removing None
        sheet_obj = SheetDict(values)
        setattr(
            sheet_obj, "name", self._getVariable(sheet)
        )  # insert a sheet's name by removing 'info-'
        return sheet_obj

    # get table sheet object
    def _getTable(self, table):
        values = list(self.wb[table].values)
        values = [x for x in values if self._notNone(x)]  # clean data by removing None
        table_obj = TableList(values)
        setattr(
            table_obj, "name", self._getVariable(table)
        )  # insert a table's name by removing 'table-'
        return table_obj

    # get the excel's specified sheets(info and table). In this step, the 'info-' and 'table-' will be removed. The rest sheet name
    # will be used as the key of the object of SheetDict or TableList.
    def _getSheetsTables(self):
        for sn in self.sheet_names:
            self.sheets[self._getVariable(sn)] = self._getSheet(sn)
        for tn in self.table_names:
            self.tables[self._getVariable(tn)] = self._getTable(tn)

    # get sheet object from the dict of self.sheets by specific sheet name,and variable list. default None means all variables. This name is variable name, without info-
    def getSheet(self, sheet_name, variables=None):
        sheet = self.sheets.get(sheet_name)
        if sheet:
            if variables:
                data = {k: v for k, v in sheet.data.items() if k in variables}
                sheet.data = data
                return sheet
            else:
                return sheet
        elif sheet == None:
            return None
        elif len(sheet) == 0:
            return sheet
        else:
            raise ValueError(
                f"Sheet info-{sheet_name} is not existed in {self.excel_name}. Or it was not referrenced in your pydantic models"
            )

    # get sheet object by specific sheet name. This name is variable name, without table-
    def getTable(self, table_name, variables=None):
        table = self.tables.get(table_name)
        if table is not None:
            if variables:
                temp_table = copy.deepcopy(table)
                if len(table.data) > 0:
                    for index, data in enumerate(table.data):
                        for k in data.__dict__:  # check every TableNode object
                            if k not in variables:  # if not in variable list
                                delattr(
                                    temp_table.data[index], k
                                )  # delete from returning obj
                temp_table.column_variables = variables  # update the table variables
                temp_table.column_titles = OrderedDict(
                    {k: v for k, v in table.column_titles.items() if k in variables}
                )  # update the table titles
                return temp_table
            else:
                return table
        else:
            raise ValueError(
                f"Table table-{table_name} is not existed in {self.excel_name}. Or it was not referrenced in your pydantic models"
            )

    # get sheet dict by specific sheet name. This name is variable name, without info-
    def getSheetData(self, sheet_name):
        # sheet_data=OrderedDict()
        if self.sheets.get(sheet_name):
            sheet = self.sheets.get(sheet_name)
            # get key(variable) and value from object
            sheet_data = {k: v.value for k, v in sheet.data.items()}
        elif len(self.sheets.get(sheet_name).data) == 0:
            sheet_data = {}
        else:
            raise ValueError(
                f"Sheet info-{sheet_name} is not existed in {self.excel_name}. Or it was not referrenced in your pydantic models"
            )

        return sheet_data

    # get table list by specific sheet name. This name is variable name, without table-
    def getTableData(self, table_name):
        table = self.tables.get(table_name, None)
        # else:
        #     raise ValueError(f"Table table-{table_name} is not existed in {self.excel_name}. Or it was not referrenced in your pydantic models")
        # if table is not existed or data is [], will return []
        if table:
            table_data = [td.__dict__ for td in table.data]
            return table_data
        return []

    # get all info-sheets and table-sheets data
    @property
    def dict(self):
        the_data = {
            **{sheet: self.getSheetData(sheet) for sheet in self.sheets.keys()},
            **{table: self.getTableData(table) for table in self.tables.keys()},
        }
        return the_data

    @property
    def json(self):
        return json.dumps(self.dict, indent=4, cls=DateEncoder)

    @property
    def plain_dict(self):
        plain_json = json.dumps(self.dict, indent=4, cls=DateEncoder)
        return json.loads(plain_json)

    """
    Below is write back to excel  mothods
    """

    def cellStyle(self, cell, format):
        cell.font = format.get("font")
        cell.border = format.get("border")
        cell.fill = format.get("fill")
        cell.alignment = format.get("alignment")

    def _formatCells(self, ws, columns):
        for col in range(1, columns + 1):
            cell = ws.cell(1, col)
            self.cellStyle(cell, style.get("TITLE_FORMAT"))

    def makeInfoSheet(self, sheet: SheetDict):
        # create a new info sheet in new workbook
        sheet_name = "info-" + sheet.name
        new_ws = self.new_wb.create_sheet(sheet_name)
        rows = []
        rows.append([sheet.sheet_title])
        rows.append(list(sheet.column_titles))
        # not use titled variables
        # titled_variables=[ x.title() for x in sheet.column_variables]
        rows.append(sheet.column_variables)
        for new_row in sheet.data.values():
            rows.append(list(new_row.__dict__.values()))
        for row in rows:
            new_ws.append(row)
        max_column = get_column_letter(len(sheet.column_titles))
        new_ws.merge_cells(f"A1:{max_column}1")

    def makeTableSheet(self, table: TableList):
        # create a new table sheet in new workbook
        table_name = "table-" + table.name
        new_ws = self.new_wb.create_sheet(table_name)
        rows = []
        rows.append([table.sheet_title])
        variables = table.column_variables
        titles = len(variables) > 0 and [table.column_titles.get(v) for v in variables]
        rows.append(titles)
        rows.append(variables)
        for new_row in table.data:
            rows.append(list(new_row.__dict__.values()))
        for row in rows:
            row and new_ws.append(row)
        max_column = get_column_letter(len(table.column_titles))
        new_ws.merge_cells(f"A1:{max_column}1")

    def makeExcel(self, file_name, protection=False, sheets=None, tables=None):
        # 创建新的excel文件
        self.new_wb = Workbook()

        self.new_wb.views = [
            BookView(xWindow=0, yWindow=0, windowWidth=18140, windowHeight=15540)
        ]

        # handle info sheets. if None, including all sheets/tables, if [], including nothing. if has sheets/tables, including those sheets/tables
        sheets = sheets if sheets is not None else self.sheets.values()
        tables = tables if tables is not None else self.tables.values()
        [self.makeInfoSheet(sheet) for sheet in sheets if sheet is not None]
        # handle table sheets
        [self.makeTableSheet(table) for table in tables if table is not None]
        # 保存为新文件
        if len(self.sheets) > 0 or len(self.tables) > 0:
            self.new_wb.remove(
                self.new_wb["Sheet"]
            )  # exists only when niether info nor table has sheet
        # old version, removed
        # f=Format(self.new_wb)
        # f.setAll()
        # if protection:
        #     f.protectAll()
        self.new_wb.save(file_name)  # 先保存
        self.formatExcel(file_name, protection)  # format这个文件

    def formatExcel(self, output_file_name, protection=True):
        input_excel_obj = Excel(output_file_name)  # get excel obj for format
        ew = ExcelWritter(input_excel_obj, output_file_name, language=self.language)
        if protection:
            password = input("please input password:")
            ew.create(protection, password)
        else:
            ew.create(protection=False)
